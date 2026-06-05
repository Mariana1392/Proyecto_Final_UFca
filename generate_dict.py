import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define the file path
output_path = r"Diccionario_Datos_UFCA_Actualizado.xlsx"

# Dictionary to hold the data for each table
data_dict = {}

# 1. roles
data_dict['roles'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador único del rol."],
    ["nombre", "", "character varying", "255", "", "NO", "SI", "SI", "Nombre descriptivo del rol de usuario (ej. 'admin', 'asociado')."],
    ["descripcion", "", "text", "MAX", "", "SI", "", "", "Explicación detallada del alcance y responsabilidades del rol."],
    ["activo", "", "boolean", "", "true", "SI", "", "", "Indica si el rol está habilitado para ser asignado."],
    ["es_sistema", "", "boolean", "", "false", "SI", "", "", "Si es true, indica que es un rol base no editable por el usuario."],
    ["created_at", "", "timestamp with time zone", "", "now()", "SI", "", "", "Fecha y hora en la que se creó el registro del rol."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "SI", "", "", "Fecha y hora de la última actualización del registro."],
    ["label", "", "text", "MAX", "", "SI", "", "", "Etiqueta en texto simple para visualización amigable en la UI."]
]

# 2. usuarios
data_dict['usuarios'] = [
    ["id", "PK", "uuid", "36", "", "NO", "SI", "SI", "Identificador del usuario (FK a auth.users de Supabase)."],
    ["rol_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea que asocia al usuario con un rol (roles.id)."],
    ["nombre", "", "character varying", "255", "", "NO", "", "", "Nombre completo del usuario."],
    ["email", "", "character varying", "255", "", "NO", "SI", "SI", "Correo electrónico único para inicio de sesión y notificaciones."],
    ["activo", "", "boolean", "", "true", "SI", "", "", "Estado del usuario (activo/inactivo)."],
    ["created_at", "", "timestamp with time zone", "", "now()", "SI", "", "", "Fecha de creación del registro de usuario."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "SI", "", "", "Fecha de última edición de los datos del usuario."],
    ["username", "", "text", "MAX", "", "SI", "", "", "Apodo o nombre corto identificador de usuario."],
    ["ultimo_acceso", "", "timestamp with time zone", "", "", "SI", "", "", "Registro del último inicio de sesión del usuario."],
    ["telefono", "", "character varying", "20", "''", "SI", "", "", "Teléfono de contacto personal del usuario."],
    ["direccion", "", "character varying", "255", "''", "SI", "", "", "Dirección de residencia física del usuario."],
    ["cedula", "", "text", "MAX", "", "SI", "", "SI", "Cédula o documento de identificación oficial del usuario."],
    ["fecha_ingreso", "", "date", "", "", "SI", "", "", "Fecha oficial en que el usuario ingresa a la cooperativa."],
    ["referido_por_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea de autoreferencia (usuarios.id) para referidos."],
    ["estado_cuenta", "", "text", "MAX", "'activo'", "NO", "", "", "Estado de cuenta ('activo', 'inactivo', 'suspendido')."]
]

# 3. permisos
data_dict['permisos'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador único del permiso en el sistema."],
    ["clave", "", "text", "MAX", "", "NO", "SI", "SI", "Clave técnica única del permiso (ej. 'usuarios.editar')."],
    ["label", "", "text", "MAX", "", "NO", "", "", "Nombre legible para las interfaces de administración de roles."],
    ["descripcion", "", "text", "MAX", "", "SI", "", "", "Explicación detallada del acceso que otorga este permiso."],
    ["grupo", "", "text", "MAX", "", "NO", "", "", "Grupo al que pertenece el permiso ('admin', 'asociado', 'usuario')."],
    ["activo", "", "boolean", "", "true", "NO", "", "", "Determina si el permiso está activo en el sistema."],
    ["created_at", "", "timestamp with time zone", "", "now()", "SI", "", "", "Fecha de creación del permiso en la plataforma."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de la última modificación del permiso."]
]

# 4. rol_permisos
data_dict['rol_permisos'] = [
    ["rol_id", "PK", "uuid", "36", "", "NO", "", "SI", "Llave foránea del rol (roles.id) - Parte 1 de la PK compuesta."],
    ["permiso_clave", "PK", "text", "MAX", "", "NO", "", "SI", "Clave técnica del permiso (permisos.clave) - Parte 2 de la PK."],
    ["asignado_en", "", "timestamp with time zone", "", "now()", "SI", "", "", "Fecha y hora en la que se le asignó el permiso a dicho rol."],
    ["activo", "", "boolean", "", "true", "NO", "", "", "Indica si la asignación está vigente actualmente."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de creación del registro."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de modificación del registro."]
]

# 5. configuracion
data_dict['configuracion'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador de la variable de configuración."],
    ["clave", "", "text", "MAX", "", "NO", "SI", "SI", "Clave única del parámetro operativo (ej. 'tasa_vivienda')."],
    ["valor", "", "text", "MAX", "", "NO", "", "", "Valor parametrizado guardado en formato de texto."],
    ["descripcion", "", "text", "MAX", "", "SI", "", "", "Breve descripción para entender qué regla controla el parámetro."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de creación de la variable."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de última edición de la variable de configuración."]
]

# 6. periodos
data_dict['periodos'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador único del periodo financiero."],
    ["fecha_inicio", "", "date", "", "", "NO", "SI", "SI", "Fecha de inicio del año o ciclo fiscal."],
    ["fecha_fin", "", "date", "", "", "NO", "", "", "Fecha de finalización del ciclo financiero."],
    ["estado", "", "text", "MAX", "'activo'", "NO", "", "", "Estado de contabilidad ('activo', 'cerrado')."],
    ["fecha_cierre", "", "timestamp with time zone", "", "", "SI", "", "", "Fecha y hora exacta en que se ejecutó el cierre del periodo."],
    ["cerrado_por", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del administrador que realizó el cierre (usuarios.id)."],
    ["utilidad_total", "", "numeric", "", "0", "NO", "", "", "Suma neta de las utilidades recolectadas durante el periodo."],
    ["utilidad_por_asociado", "", "numeric", "", "", "SI", "", "", "Excedentes distribuidos equitativamente por asociado."],
    ["num_asociados_activos", "", "integer", "", "", "SI", "", "", "Cantidad de asociados activos al momento del cierre."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de creación del periodo."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de última edición del periodo."],
    ["nombre", "", "text", "MAX", "calculated", "SI", "", "", "Nombre descriptivo dinámico del periodo (ej. 'Período 2026/2027')."]
]

# 7. solicitudes_asociados
data_dict['solicitudes_asociados'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador de la postulación de afiliación."],
    ["nombres", "", "text", "MAX", "", "NO", "", "", "Nombres del aspirante a asociado."],
    ["apellidos", "", "text", "MAX", "", "NO", "", "", "Apellidos del aspirante a asociado."],
    ["cedula", "", "text", "MAX", "", "NO", "SI", "SI", "Cédula única del aspirante para evitar múltiples afiliaciones."],
    ["tipo_identificacion", "", "text", "MAX", "", "SI", "", "", "Tipo de documento (cédula, pasaporte, cédula de extranjería)."],
    ["telefono", "", "text", "MAX", "", "SI", "", "", "Número telefónico de contacto del aspirante."],
    ["email", "", "text", "MAX", "", "SI", "", "", "Dirección de correo del aspirante."],
    ["direccion", "", "text", "MAX", "", "SI", "", "", "Dirección domiciliaria del aspirante."],
    ["ocupacion", "", "text", "MAX", "", "SI", "", "", "Ocupación, profesión o actividad económica."],
    ["ingreso_mensual", "", "numeric", "", "", "SI", "", "", "Ingresos mensuales aproximados declarados."],
    ["motivacion", "", "text", "MAX", "", "SI", "", "", "Breve texto sobre el motivo del ingreso."],
    ["estado", "", "text", "MAX", "'pendiente'", "NO", "", "", "Estado ('pendiente', 'aprobada', 'rechazada', 'pendiente_activacion')."],
    ["documentos", "", "ARRAY", "", "", "SI", "", "", "Colección de URLs que contienen la documentación adjunta."],
    ["observaciones", "", "text", "MAX", "", "SI", "", "", "Anotaciones de la mesa de control o comité."],
    ["fecha_solicitud", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha y hora en la que se envió la solicitud."],
    ["fecha_resolucion", "", "timestamp with time zone", "", "", "SI", "", "", "Fecha en que el comité aprobó o rechazó la afiliación."],
    ["fecha_activacion", "", "date", "", "", "SI", "", "", "Fecha de la activación formal (posterior a la primera cuota)."],
    ["usuario_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del usuario creado vinculado (usuarios.id)."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de inserción de la fila."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de última edición."],
    ["resuelto_por", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del usuario evaluador que emitió resolución (usuarios.id)."],
    ["monto_ahorro_propuesto", "", "numeric", "", "", "SI", "", "", "Cuota mensual sugerida por el usuario para su ahorro permanente."],
    ["aprobado_por", "", "uuid", "36", "", "SI", "", "SI", "Administrador que aprueba definitivamente el perfil (usuarios.id)."],
    ["recordatorio_enviado_at", "", "timestamp with time zone", "", "", "SI", "", "", "Fecha del último correo de aviso enviado."]
]

# 8. comite_evaluador
data_dict['comite_evaluador'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador del registro de comité."],
    ["solicitud_asociado_id", "", "uuid", "36", "", "NO", "SI", "SI", "FK de la solicitud evaluada (solicitudes_asociados.id)."],
    ["evaluador_id", "", "uuid", "36", "", "SI", "", "SI", "FK del evaluador asignado (usuarios.id)."],
    ["verificaciones", "", "jsonb", "", "'{\"ingresos\"...'", "NO", "", "", "Estado de validación documental en JSON."],
    ["score_credito", "", "integer", "", "70", "NO", "", "", "Puntaje estimado de viabilidad de crédito/afiliación (0-100)."],
    ["comentarios", "", "text", "MAX", "", "SI", "", "", "Comentarios del dictamen del comité."],
    ["decision", "", "text", "MAX", "'en_evaluacion'", "NO", "", "", "Decisión tomada ('en_evaluacion', 'aprobado', 'rechazado')."],
    ["observacion", "", "text", "MAX", "", "SI", "", "", "Observaciones de cierre y excepciones."],
    ["fecha", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de emisión de la deliberación del comité."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de creación del log."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de última edición."]
]

# 9. creditos
data_dict['creditos'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador único del préstamo."],
    ["asociado_id", "", "uuid", "36", "", "NO", "", "SI", "Llave foránea del asociado deudor (usuarios.id)."],
    ["periodo_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del periodo financiero activo (periodos.id)."],
    ["tipo", "", "text", "MAX", "", "NO", "", "", "Línea de crédito ('libre_inversion', 'educacion', 'vivienda', 'calamidad')."],
    ["monto", "", "numeric", "", "", "NO", "", "", "Valor nominal solicitado/prestado."],
    ["plazo_meses", "", "integer", "", "", "NO", "", "", "Número de cuotas mensuales pactadas para pago (1-12)."],
    ["tasa_interes", "", "numeric", "", "", "NO", "", "", "Tasa de interés remuneratoria aplicada al crédito."],
    ["tasa_mora", "", "numeric", "", "", "SI", "", "", "Tasa de interés penal de mora en caso de incumplimiento."],
    ["cuota_mensual", "", "numeric", "", "", "NO", "", "", "Valor periódico amortizado calculado de la cuota."],
    ["saldo", "", "numeric", "", "", "NO", "", "", "Saldo insoluto pendiente de pago actual."],
    ["estado", "", "text", "MAX", "'pendiente'", "NO", "", "", "Estado ('pendiente', 'en_revision', 'desembolsado', 'activo', 'en_mora'...)."],
    ["fecha_desembolso", "", "date", "", "", "SI", "", "", "Fecha en la cual se entregó el capital prestado."],
    ["fecha_primera_cuota", "", "date", "", "", "SI", "", "", "Fecha en que vence la primera cuota proyectada."],
    ["fecha_ultima_cuota", "", "date", "", "", "SI", "", "", "Fecha límite en que vence la última cuota pactada."],
    ["fecha_estado_cambio", "", "timestamp with time zone", "", "", "SI", "", "", "Timestamp del último cambio de estado del crédito."],
    ["motivo_estado_cambio", "", "text", "MAX", "", "SI", "", "", "Justificación escrita del cambio de estado."],
    ["url_comprobante_solicitud", "", "text", "MAX", "", "SI", "", "", "Comprobante físico o digital de radicación."],
    ["anulado", "", "boolean", "", "false", "NO", "", "", "Identifica si el préstamo fue cancelado/invalidado administrativamente."],
    ["motivo_anulacion", "", "text", "MAX", "", "SI", "", "", "Razón por la cual se procedió a la anulación."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de creación del préstamo."],
    ["observaciones", "", "text", "MAX", "", "SI", "", "", "Notas generales internas del crédito."],
    ["tipo_interes", "", "text", "MAX", "'compuesto'", "SI", "", "", "Modalidad de cobro de interés ('simple', 'compuesto')."],
    ["anulado_por", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del administrador que anuló (usuarios.id)."],
    ["anulado_en", "", "timestamp with time zone", "", "", "SI", "", "", "Fecha en la que ocurrió la anulación."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "SI", "", "", "Fecha de última edición del registro."]
]

# 10. cuotas_credito
data_dict['cuotas_credito'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador de la cuota del préstamo."],
    ["credito_id", "", "uuid", "36", "", "NO", "", "SI", "Llave foránea del préstamo asociado (creditos.id)."],
    ["num_cuota", "", "integer", "", "", "NO", "", "", "Número correlativo de la cuota del crédito."],
    ["fecha_vencimiento", "", "date", "", "", "NO", "", "", "Fecha límite en que debe pagarse esta cuota."],
    ["capital", "", "numeric", "", "", "NO", "", "", "Fracción monetaria que amortiza el saldo del capital prestado."],
    ["interes", "", "numeric", "", "", "NO", "", "", "Fracción de cobro correspondiente al interés ordinario de la cuota."],
    ["cuota_total", "", "numeric", "", "", "NO", "", "", "Sumatoria total de la cuota (capital + interés)."],
    ["saldo_inicial", "", "numeric", "", "", "NO", "", "", "Saldo de capital al inicio del mes."],
    ["saldo_final", "", "numeric", "", "", "NO", "", "", "Saldo de capital restante luego de pagar la cuota."],
    ["estado", "", "text", "MAX", "'pendiente'", "NO", "", "", "Estado de la cuota ('pendiente', 'pagada', 'mora', 'abono_aplicado')."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de creación del registro."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de modificación del registro."]
]

# 11. liquidaciones
data_dict['liquidaciones'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador único del registro de liquidación."],
    ["asociado_id", "", "uuid", "36", "", "NO", "", "SI", "Llave foránea del asociado en liquidación (usuarios.id)."],
    ["periodo_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del periodo financiero de corte (periodos.id)."],
    ["usuario_id", "", "uuid", "36", "", "SI", "", "SI", "FK del funcionario/admin liquidador (usuarios.id)."],
    ["tipo", "", "text", "MAX", "", "NO", "", "", "Causa ('retiro', 'cierre_anual', 'fallecimiento', 'expulsion', 'otro')."],
    ["total_ahorro_permanente", "", "numeric", "", "0", "NO", "", "", "Monto acumulado devuelto de ahorro permanente."],
    ["total_ahorro_voluntario", "", "numeric", "", "0", "NO", "", "", "Monto acumulado devuelto de ahorro voluntario."],
    ["total_deudas_credito", "", "numeric", "", "0", "NO", "", "", "Deudas de créditos activos descontadas de la liquidación."],
    ["utilidades", "", "numeric", "", "0", "NO", "", "", "Utilidades/excedentes distribuidos en favor del asociado."],
    ["monto_neto", "", "numeric", "", "0", "NO", "", "", "Monto total neto devuelto (Ahorros + Excedentes - Deudas)."],
    ["detalle", "", "jsonb", "", "", "SI", "", "", "Desglose pormenorizado en JSON."],
    ["observaciones", "", "text", "MAX", "", "SI", "", "", "Anotaciones e incidencias de la liquidación."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha del cálculo."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de última edición."],
    ["fecha", "", "date", "", "", "SI", "", "", "Fecha oficial en que se formalizó la liquidación."],
    ["estado", "", "text", "MAX", "'En proceso'", "SI", "", "", "Estado del trámite administrativo (ej. 'Completada', 'Anulada')."],
    ["fecha_corte", "", "date", "", "", "SI", "", "", "Fecha para el cálculo de saldos consolidados."],
    ["fecha_liquidacion", "", "date", "", "", "SI", "", "", "Fecha de desembolso del dinero neto de la liquidación."],
    ["anulado", "", "boolean", "", "false", "SI", "", "", "Indica si la liquidación fue invalidada."],
    ["justificacion_anulacion", "", "text", "MAX", "", "SI", "", "", "Explicación del porqué de la anulación."],
    ["anulado_por", "", "text", "MAX", "", "SI", "", "", "Nombre del administrador ejecutor de la anulación."],
    ["anulado_en", "", "timestamp with time zone", "", "", "SI", "", "", "Timestamp de la anulación."],
    ["monto_total", "", "numeric", "", "0", "NO", "", "", "Consolidado total bruto de los activos calculados."]
]

# 12. distribuciones_utilidades
data_dict['distribuciones_utilidades'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador de la distribución de utilidades."],
    ["periodo_id", "", "uuid", "36", "", "NO", "", "SI", "Llave foránea del periodo contable (periodos.id)."],
    ["asociado_id", "", "uuid", "36", "", "NO", "", "SI", "Llave foránea del asociado receptor (usuarios.id)."],
    ["utilidad_total_periodo", "", "numeric", "", "", "NO", "", "", "Monto acumulado bruto de utilidad cooperativa en el periodo."],
    ["num_asociados", "", "integer", "", "", "NO", "", "", "Cantidad de asociados receptores de la utilidad."],
    ["valor_por_asociado", "", "numeric", "", "", "NO", "", "", "Porción individual distribuida al asociado."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha del asiento."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de última edición."]
]

# 13. excepciones
data_dict['excepciones'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador único de la excepción."],
    ["asociado_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea opcional del asociado relacionado (usuarios.id)."],
    ["credito_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea opcional del crédito relacionado (creditos.id)."],
    ["tipo", "", "text", "MAX", "", "NO", "", "", "Código técnico de la regla violada (ej. 'retiro_con_deudas')."],
    ["descripcion", "", "text", "MAX", "", "NO", "", "", "Explicación detallada del motivo por el cual se pide la excepción."],
    ["estado", "", "text", "MAX", "'pendiente'", "NO", "", "", "Estado de resolución administrativa ('pendiente', 'aprobada', 'rechazada')."],
    ["resuelto_por", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del administrador aprobador (usuarios.id)."],
    ["fecha_resolucion", "", "timestamp with time zone", "", "", "SI", "", "", "Fecha y hora en la que se aprobó o rechazó."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de radicación de la excepción."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de última edición de la excepción."]
]

# 14. notificaciones
data_dict['notificaciones'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador único de la alerta."],
    ["usuario_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del destinatario principal (usuarios.id)."],
    ["asociado_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del asociado relacionado (usuarios.id)."],
    ["tipo", "", "text", "MAX", "", "NO", "", "", "Tipo ('credito_pendiente', 'credito_activo', 'pago_registrado'...)."],
    ["titulo", "", "text", "MAX", "", "NO", "", "", "Título que describe la notificación (ej. 'Crédito Aprobado')."],
    ["mensaje", "", "text", "MAX", "", "NO", "", "", "Mensaje descriptivo con detalles de la notificación."],
    ["leida", "", "boolean", "", "false", "NO", "", "", "Bandera lógica para saber si el destinatario la abrió."],
    ["para_admin", "", "boolean", "", "false", "NO", "", "", "Si es true, notifica a todos los administradores del sistema."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de generación de la alerta."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de última edición."]
]

# 15. auditoria
data_dict['auditoria'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador único de auditoría."],
    ["usuario_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del ejecutor de la operación (usuarios.id)."],
    ["asociado_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del asociado que se modificó (usuarios.id)."],
    ["tabla", "", "text", "MAX", "", "NO", "", "", "Nombre de la tabla afectada (ej. 'creditos', 'transacciones')."],
    ["registro_id", "", "uuid", "36", "", "SI", "", "", "ID exacto de la fila afectada en la tabla mencionada."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha y hora exacta de la transacción."],
    ["operacion", "", "character varying", "255", "", "SI", "", "", "Tipo de DML ejecutado ('INSERT', 'UPDATE', 'DELETE')."],
    ["datos_antes", "", "jsonb", "", "", "SI", "", "", "Snapshot JSON con el estado original de la fila."],
    ["datos_despues", "", "jsonb", "", "", "SI", "", "", "Snapshot JSON con el estado de la fila post-operación."],
    ["accion", "", "character varying", "255", "", "SI", "", "", "Clave descriptiva de la acción de negocio (ej. 'APORTE_ANULADO')."]
]

# 16. referidos
data_dict['referidos'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador único del referido."],
    ["asociado_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del asociado originador (usuarios.id)."],
    ["nombre", "", "text", "MAX", "", "NO", "", "", "Nombre completo de la persona recomendada."],
    ["cedula", "", "text", "MAX", "", "NO", "", "", "Cédula del referido."],
    ["telefono", "", "text", "MAX", "", "SI", "", "", "Teléfono de contacto de la persona referida."],
    ["estado", "", "character varying", "255", "'activo'", "NO", "", "", "Estado ('activo', 'inactivo')."],
    ["observaciones", "", "text", "MAX", "", "SI", "", "", "Anotaciones e incidencias del referido."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de ingreso de la recomendación."],
    ["asociado_convertido_id", "", "uuid", "36", "", "SI", "", "SI", "FK del nuevo asociado creado a partir de la conversión (usuarios.id)."],
    ["fecha_conversion", "", "date", "", "", "SI", "", "", "Fecha en que el referido completó su afiliación y fue activado."]
]

# 17. credito_historial_estados
data_dict['credito_historial_estados'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador del log de estado del crédito."],
    ["credito_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del crédito correspondiente (creditos.id)."],
    ["estado_anterior", "", "text", "MAX", "", "SI", "", "", "Estado del crédito previo al cambio."],
    ["estado_nuevo", "", "text", "MAX", "", "SI", "", "", "Estado al cual transicionó el crédito."],
    ["motivo", "", "text", "MAX", "", "SI", "", "", "Justificación o notas sobre el cambio de estado."],
    ["cambiado_por", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del administrador responsable (usuarios.id)."],
    ["cambiado_en", "", "timestamp with time zone", "", "now()", "SI", "", "", "Fecha y hora en que se ejecutó la actualización."]
]

# 18. cuentas_ahorro
data_dict['cuentas_ahorro'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador de la cuenta de ahorro de asociado."],
    ["tipo", "", "text", "MAX", "", "NO", "", "", "Tipo de cuenta ('permanente', 'voluntario')."],
    ["asociado_id", "", "uuid", "36", "", "NO", "", "SI", "Llave foránea del asociado propietario (usuarios.id)."],
    ["periodo_id", "", "uuid", "36", "", "NO", "", "SI", "Llave foránea del periodo financiero activo de la cuenta (periodos.id)."],
    ["monto_ahorrado", "", "numeric", "", "0", "NO", "", "", "Saldo consolidado total actual acumulado de la cuenta."],
    ["cuota_mensual", "", "numeric", "", "", "SI", "", "", "Cuota mensual fija pactada para debitar/aportar."],
    ["fecha_retiro", "", "timestamp with time zone", "", "", "SI", "", "", "Fecha en que se liquidó y retiró el ahorro total."],
    ["monto_al_cierre", "", "numeric", "", "", "SI", "", "", "Saldo definitivo reportado al cierre del ejercicio anual."],
    ["estado", "", "text", "MAX", "'activo'", "NO", "", "", "Estado de la cuenta ('activo', 'cerrado', 'suspendido', 'retirado')."],
    ["fecha_cierre", "", "timestamp with time zone", "", "", "SI", "", "", "Fecha y hora exactas del cierre de la cuenta de ahorro."],
    ["anulado", "", "boolean", "", "false", "NO", "", "", "Marca lógica si la cuenta se anuló administrativamente."],
    ["anulado_por", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del administrador que anuló (usuarios.id)."],
    ["anulado_en", "", "timestamp with time zone", "", "", "SI", "", "", "Fecha y hora exactas de la anulación."],
    ["motivo_anulacion", "", "text", "MAX", "", "SI", "", "", "Razón justificada de la anulación de la cuenta."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de apertura de la cuenta de ahorro."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de última edición."],
    ["observaciones", "", "text", "MAX", "", "SI", "", "", "Comentarios u observaciones de control sobre la cuenta."],
    ["cedula", "", "text", "MAX", "", "SI", "", "", "Cédula del asociado vinculada directamente (para consultas rápidas)."]
]

# 19. transacciones
data_dict['transacciones'] = [
    ["id", "PK", "uuid", "36", "gen_random_uuid()", "NO", "SI", "SI", "Identificador único del movimiento financiero."],
    ["tipo", "", "text", "MAX", "", "NO", "", "", "Tipo de movimiento ('aporte_permanente', 'aporte_voluntario', 'pago_credito'...)."],
    ["asociado_id", "", "uuid", "36", "", "NO", "", "SI", "Llave foránea del asociado dueño del movimiento (usuarios.id)."],
    ["registrado_por", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del cajero o administrador responsable (usuarios.id)."],
    ["ahorro_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea a la cuenta de ahorros afectada (cuentas_ahorro.id)."],
    ["credito_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea al crédito amortizado (creditos.id)."],
    ["cuota_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea a la cuota específica pagada (cuotas_credito.id)."],
    ["periodo_id", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del periodo financiero activo (periodos.id)."],
    ["monto", "", "numeric", "", "", "NO", "", "", "Valor total monetario operado en la transacción."],
    ["capital", "", "numeric", "", "0", "NO", "", "", "Porción del monto destinada a amortizar el capital."],
    ["interes", "", "numeric", "", "0", "NO", "", "", "Porción del monto destinada a amortizar intereses."],
    ["monto_mora", "", "numeric", "", "0", "NO", "", "", "Porción del monto cobrada en calidad de interés por mora."],
    ["dias_mora", "", "integer", "", "0", "NO", "", "", "Días de mora contabilizados en el crédito al momento de pagar."],
    ["saldo_antes", "", "numeric", "", "", "SI", "", "", "Saldo de la cuenta o del préstamo antes de la transacción."],
    ["saldo_despues", "", "numeric", "", "", "SI", "", "", "Saldo de la cuenta o del préstamo resultante de la transacción."],
    ["mes_correspondiente", "", "date", "", "", "SI", "", "", "Mes del año al cual imputa el aporte del asociado."],
    ["fecha_pago", "", "date", "", "", "NO", "", "", "Fecha real en la que el usuario efectuó el pago."],
    ["metodo_pago", "", "text", "MAX", "", "SI", "", "", "Medio de pago utilizado (ej. 'transferencia', 'efectivo')."],
    ["url_comprobante", "", "text", "MAX", "", "SI", "", "", "Enlace al archivo de soporte/comprobante en el Storage."],
    ["observacion", "", "text", "MAX", "", "SI", "", "", "Notas u observaciones del registro."],
    ["anulado", "", "boolean", "", "false", "NO", "", "", "Bandera que indica si la transacción fue reversada/anulada."],
    ["anulado_por", "", "uuid", "36", "", "SI", "", "SI", "Llave foránea del administrador que anuló (usuarios.id)."],
    ["anulado_en", "", "timestamp with time zone", "", "", "SI", "", "", "Timestamp exacto de la anulación."],
    ["motivo_anulacion", "", "text", "MAX", "", "SI", "", "", "Justificación de la anulación."],
    ["created_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de registro inicial."],
    ["updated_at", "", "timestamp with time zone", "", "now()", "NO", "", "", "Fecha de última edición del registro."]
]

# Write to Excel
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    for sheet_name, rows in data_dict.items():
        # Create DataFrame
        df = pd.DataFrame(rows, columns=["Campo", "PK", "Tipo", "Tamaño", "Default", "Null", "Único", "Indexado", "Observaciones"])
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Style the sheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Style header
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # Apply style to headers
        for col_idx in range(1, 10):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = thin_border
            
        # Style rows and center specific columns
        row_font = Font(name="Calibri", size=11, color="000000")
        for r_idx in range(2, len(rows) + 2):
            for col_idx in range(1, 10):
                cell = worksheet.cell(row=r_idx, column=col_idx)
                cell.font = row_font
                cell.border = thin_border
                
                # Alignments
                if col_idx in [2, 4, 6, 7, 8]:  # PK, Tamaño, Null, Único, Indexado
                    cell.alignment = alignment_center
                else:
                    cell.alignment = alignment_left
                    
        # Adjust row height
        worksheet.row_dimensions[1].height = 28
        for r_idx in range(2, len(rows) + 2):
            worksheet.row_dimensions[r_idx].height = 20
            
        # Adjust column widths
        column_widths = {
            'A': 22,  # Campo
            'B': 8,   # PK
            'C': 22,  # Tipo
            'D': 10,  # Tamaño
            'E': 20,  # Default
            'F': 8,   # Null
            'G': 8,   # Único
            'H': 10,  # Indexado
            'I': 70   # Observaciones
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

print("Excel file generated successfully!")
