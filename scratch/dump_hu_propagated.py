import openpyxl
import json

def dump_excel():
    file_path = r"C:\Users\dairi\Downloads\Historias de usuario UFCA (1).xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_name = "HU + Criterios Aceptación"
    ws = wb[sheet_name]
    
    headers = [
        "Proceso", "Subproceso", "Epica", "Yo_como", "Rol", 
        "Deseo_necesito_quiero", "Objetivo", "Para_poder", 
        "Beneficio", "Codigo_HU", "Redaccion_HU", 
        "Codigo_CA", "Criterio", "Estado"
    ]
    
    current_values = {h: None for h in headers[:11]}
    results = []
    
    for r in range(4, ws.max_row + 1):
        # Read columns 1 to 14
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
        
        # Check if row is completely empty
        if all(v is None for v in row_vals):
            continue
            
        # Propagate columns 1 to 11
        for i, h in enumerate(headers[:11]):
            val = row_vals[i]
            if val is not None and str(val).strip() != "":
                current_values[h] = str(val).strip()
                
        # Get columns 12 to 14
        cod_ca = row_vals[11]
        criterio = row_vals[12]
        estado = row_vals[13]
        
        # Strip string values if they exist
        cod_ca = str(cod_ca).strip() if cod_ca is not None else None
        criterio = str(criterio).strip() if criterio is not None else None
        estado = str(estado).strip() if estado is not None else None
        
        # Only include if we have a valid acceptance criterion
        if cod_ca or criterio:
            item = {
                "row_number": r,
                "codigo_ca": cod_ca,
                "criterio": criterio,
                "estado_ca": estado
            }
            # Add the propagated story info
            for h in headers[:11]:
                item[h.lower()] = current_values[h]
            results.append(item)
            
    print(f"Total acceptance criteria found: {len(results)}")
    
    with open("scratch/hu_criterios_propagated.json", "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print("Saved to scratch/hu_criterios_propagated.json")

if __name__ == "__main__":
    dump_excel()
