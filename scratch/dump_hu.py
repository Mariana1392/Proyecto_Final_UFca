import openpyxl
import json

def dump_excel():
    file_path = r"C:\Users\dairi\Downloads\Historias de usuario UFCA (1).xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_name = "HU + Criterios Aceptación"
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
        return
    
    ws = wb[sheet_name]
    
    rows_data = []
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    # filter out trailing Nones in headers
    while headers and headers[-1] is None:
        headers.pop()
        
    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        # Skip completely empty rows
        if any(v is not None for v in row_vals):
            rows_data.append({
                "row": r,
                "values": row_vals
            })
            
    output = {
        "headers": headers,
        "rows": rows_data
    }
    
    with open("scratch/hu_criterios_full.json", "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print("Dumped successfully to scratch/hu_criterios_full.json")

if __name__ == "__main__":
    dump_excel()
