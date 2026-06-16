import openpyxl
import json
import sys

def read_excel():
    file_path = r"C:\Users\dairi\Downloads\Historias de usuario UFCA (1).xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_name = "HU + Criterios Aceptación"
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
        return
    
    ws = wb[sheet_name]
    print(f"Reading sheet '{sheet_name}' (max_row={ws.max_row}, max_column={ws.max_column})")
    
    # We want to print all rows
    rows_data = []
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        # Skip completely empty rows
        if any(v is not None for v in row_vals):
            rows_data.append((r, row_vals))
            
    print(f"Total non-empty rows: {len(rows_data)}")
    for r_idx, vals in rows_data:
        # Let's filter out None elements at the end to make it cleaner
        while vals and vals[-1] is None:
            vals.pop()
        print(f"Row {r_idx:03d}: {vals}")

if __name__ == "__main__":
    read_excel()
