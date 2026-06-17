import openpyxl

def inspect():
    file_path = r"C:\Users\dairi\Downloads\Historias de usuario UFCA (1).xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_name = "HU + Criterios Aceptación"
    ws = wb[sheet_name]
    
    print("Max Row:", ws.max_row)
    print("Max Column:", ws.max_column)
    
    # Print first 10 rows completely up to column 20
    for r in range(1, 15):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 20)]
        print(f"Row {r:02d}: {vals}")

if __name__ == "__main__":
    inspect()
