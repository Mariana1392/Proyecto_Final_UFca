import openpyxl

def analyze(file_path):
    print(f"\nAnalyzing {file_path}...")
    try:
        wb = openpyxl.load_workbook(file_path)
        print("Sheets in workbook:", wb.sheetnames)
        for name in wb.sheetnames:
            ws = wb[name]
            print(f"- Sheet '{name}': max_row={ws.max_row}, max_column={ws.max_column}")
            # print first 5 rows of data
            for r in range(1, min(6, ws.max_row + 1)):
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(6, ws.max_column + 1))]
                print(f"  Row {r}: {row_vals}")
    except Exception as e:
        print("Error analyzing:", e)

analyze(r"C:\Users\dairi\Downloads\Reporte_Consolidado_Sistema_UFCA_2026-06-14.xlsx")
analyze(r"C:\Users\dairi\Downloads\Reporte_Consolidado_Sistema_UFCA_2026-06-14 (1).xlsx")
